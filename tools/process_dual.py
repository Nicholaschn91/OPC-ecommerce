#!/usr/bin/env python3
"""
通用双源关键词处理引擎（erank CSV + 西柚多ASIN xlsx）
Usage: python process_dual.py <input_dir> <spu_id> <spu_name> [category]
"""

import csv, os, sys, sqlite3, re, glob
import pandas as pd
import numpy as np
from collections import Counter
from datetime import datetime

# ──── Config ────
INPUT_DIR = sys.argv[1]
SPU_ID = sys.argv[2]
SPU_NAME = sys.argv[3]
CAT = sys.argv[4] if len(sys.argv) > 4 else "生活家居"
BASE = r"C:/Users/Administrator.DESKTOP-AHRMISP/Desktop/keywords/erank keywords"
WS = os.path.dirname(os.path.abspath(__file__))

# ──── 1. 合并 erank CSV ────
csv_files = glob.glob(os.path.join(INPUT_DIR, "*.csv"))
encodings = ['utf-8-sig', 'gbk', 'utf-16', 'latin-1']
RAW = {'kw': 0, 'comp': 1, 'views': 3, 'favs': 5, 'sales': 7, 'revs': 9}

def open_csv(fp):
    for enc in encodings:
        try:
            f = open(fp, 'r', encoding=enc)
            f.read(1)
            f.seek(0)
            return f, enc
        except:
            try:
                f.close()
            except:
                pass
    return open(fp, 'r', encoding='utf-8', errors='replace'), 'fallback'

def clean(v):
    return v.replace(',', '').replace('，', '').strip()

erank_data = {}
seen = {}
total = dup = 0

if csv_files:
    for fp in csv_files:
        f, enc = open_csv(fp)
        rdr = csv.reader(f)
        for row in rdr:
            if rdr.line_num == 1:
                continue
            total += 1
            try:
                kw = clean(row[RAW['kw']])
                comp = int(clean(row[RAW['comp']])) if clean(row[RAW['comp']]).isdigit() else 0
                views = int(clean(row[RAW['views']])) if clean(row[RAW['views']]).isdigit() else 0
                favs = int(clean(row[RAW['favs']])) if clean(row[RAW['favs']]).isdigit() else 0
                sales = int(clean(row[RAW['sales']])) if clean(row[RAW['sales']]).isdigit() else 0
                revs = int(clean(row[RAW['revs']])) if clean(row[RAW['revs']]).isdigit() else 0
            except:
                continue
            sk = kw.lower()
            if sk in seen:
                dup += 1
            else:
                sdr = round(views / (comp + 1), 4) if views else 0
                cvr = round(sales / (views + 1), 6) if views else 0
                pi = round(favs / (views + 1), 6) if views else 0
                seen[sk] = {
                    'kw': kw, 'mo_views': views, 'mo_comp': comp, 'sdr': sdr,
                    'mo_sales': sales, 'cvr': cvr, 'mo_favs': favs, 'mo_revs': revs, 'biz': pi
                }
        f.close()
    print(f'erank: {total}行 from {len(csv_files)} files, dup={dup}, unique={len(seen)}')
    erank_data = list(seen.values())
else:
    print('erank: 无CSV文件')

# ──── 2. 处理西柚 xlsx ────
xlsx_files = glob.glob(os.path.join(INPUT_DIR, "多ASIN对比结果*.xlsx"))
siyou_results = []

if xlsx_files:
    xp = xlsx_files[0]
    df = pd.read_excel(xp, sheet_name=0)
    col_map = {
        '关键词 (数据来源于西柚洞察)': 'kw', '翻译': 'trans', '词标签': 'tag',
        '相关性得分': 'rel_score', '相关性档位': 'rel_tier',
        '流量总和': 'total_traffic', '周平均关键词排名': 'wkly_rank',
        '周平均搜索量': 'wkly_search_vol', 'CPC建议竞价($)': 'cpc',
        '建议竞价范围($)': 'cpc_range',
        '点击转化率(均值)': 'click_cvr', '周平均竞争难度': 'wkly_comp',
        '竞争难度档位': 'comp_tier', '自然位滚动率': 'position_vol',
        'Top3周平均点击份额': 'top3_click_share',
        'Top3周平均转化份额': 'top3_cvr_share', 'Top3 ASIN': 'top3_asin',
        'asin数量': 'asin_count'
    }
    df.rename(columns={k: v for k, v in col_map.items() if k in df.columns}, inplace=True)

    asin_cols = [c for c in df.columns if c.endswith('-自然排名')]
    asins = sorted(set(c.split('-')[0] for c in asin_cols))
    print(f'西柚: {len(df)}行, 竞品={asins}')

    df_f = df[df.rel_score > 0].copy()
    print(f'  相关(>0): {len(df_f)}, 丢弃: {len(df) - len(df_f)}')

    df_f['mo_views'] = (df_f.wkly_search_vol * 4.3).astype(int)
    df_f['mo_comp'] = df_f.wkly_comp.fillna(0)
    df_f['mo_cvr'] = df_f.click_cvr.fillna(0)
    df_f['mo_cpc'] = df_f.cpc.fillna(0)
    df_f['biz_value'] = df_f.mo_cpc * df_f.mo_cvr

    df_f['competitor_count'] = 0
    for a in asins:
        sp_col = f'{a}-SP广告排名'
        org_col = f'{a}-自然排名'
        if sp_col in df_f.columns:
            df_f['competitor_count'] += df_f[sp_col].notna().astype(int)
        elif org_col in df_f.columns:
            df_f['competitor_count'] += df_f[org_col].notna().astype(int)

    n = len(df_f)
    vs = sorted(df_f.mo_views.tolist())
    cs = sorted(df_f.mo_comp.tolist())
    cvs = sorted(df_f.mo_cvr.tolist())
    vos = sorted(df_f.position_vol.dropna().tolist())

    def pct(arr, p):
        return arr[min(len(arr) - 1, max(0, int(len(arr) * p)))] if arr else 0

    has_cvr = df_f[df_f.mo_cvr > 0]
    cvr_n = len(has_cvr)
    cvr_vs = sorted(has_cvr.mo_cvr.tolist()) if cvr_n > 0 else [0.05]

    bm = {
        'p90v': pct(vs, .9), 'p60v': pct(vs, .6), 'p40v': pct(vs, .4),
        'p15v': pct(vs, .15), 'medv': pct(vs, .5), 'med_comp': pct(cs, .5),
        'med_cvr_eff': pct(cvr_vs, .5), 'p80_vol': pct(vos, .8)
    }
    print(f'  基准: P90v={bm["p90v"]:.0f} medCVR={bm["med_cvr_eff"]:.4f}')

    def classify_siyou(r):
        v, cvr, cpc, biz, vol = r.mo_views, r.mo_cvr, r.mo_cpc, r.biz_value, r.position_vol
        cc = r.competitor_count
        rel = r.rel_tier
        has_data = cvr > 0 or cpc > 0
        if pd.notna(rel) and rel in ('低相关', '极低相关'):
            return ('T5', '泛流量词', '跨品类弱关联')
        if pd.notna(vol) and vol >= bm['p80_vol']:
            return ('T4', '不稳定词', '排名滚动率高，不建议重投')
        if v >= bm['p90v']:
            return ('T1', '核心大词', '高流量+多竞品必争' if cc >= 2 else '高流量但有空间')
        if cc >= 2 and v >= bm['p40v']:
            return ('T2', '竞品必争', '多竞品投放+高转化验证' if (has_data and cvr >= bm['med_cvr_eff']) else '多竞品投放，实战验证')
        if has_data and v >= bm['p60v'] and cvr >= bm['med_cvr_eff']:
            return ('T2', '流量基石', '流量+转化双优')
        if v >= bm['p60v']:
            return ('T2', '流量基石', '流量中上')
        if has_data and cvr >= bm['med_cvr_eff']:
            return ('T3', '长尾引流', '转化好但流量低')
        if cc == 1 and v >= bm['p15v']:
            return ('T3', '长尾潜力', '单竞品验证，值得试探')
        if v >= bm['p15v']:
            return ('T4', '利润尖刀', '低流量待测')
        return ('T4', '利润尖刀', '尾部小词')

    for _, r in df_f.iterrows():
        tier, tag, dec = classify_siyou(r)
        siyou_results.append({
            'kw': r.kw, 'mo_views': r.mo_views, 'mo_comp': int(r.mo_comp),
            'mo_sales': 0, 'mo_favs': 0, 'mo_revs': 0,
            'sdr': round(r.mo_views / (r.mo_comp + 1), 4) if r.mo_comp else 999,
            'cvr': r.mo_cvr, 'biz': r.biz_value,
            'cpc': r.mo_cpc, 'position_vol': r.position_vol if pd.notna(r.position_vol) else 0,
            'top3_click': r.top3_click_share if pd.notna(r.top3_click_share) else 0,
            'top3_cvr': r.top3_cvr_share if pd.notna(r.top3_cvr_share) else 0,
            'comp_count': r.competitor_count, 'rel_score': r.rel_score, 'rel_tier': r.rel_tier,
            'tier': tier, 'tag': tag, 'decision': dec, 'source': '西柚'
        })
    tc = Counter(r['tier'] for r in siyou_results)
    print(f'  西柚T1-T5: {dict(tc)}')
else:
    print('西柚: 无xlsx文件')

# ──── 3. 双源融合 ────
erank_kws = {d['kw'].lower(): d for d in erank_data}
siyou_kws = {s['kw'].lower(): s for s in siyou_results}
all_kws = set(list(erank_kws.keys()) + list(siyou_kws.keys()))
merged = []

for kw in all_kws:
    ek = erank_kws.get(kw)
    sk = siyou_kws.get(kw)
    if ek and sk:
        d = {
            'kw': sk['kw'],
            'mo_views': max(ek['mo_views'], sk['mo_views']),
            'mo_comp': sk['mo_comp'] if sk['mo_comp'] > 0 else ek['mo_comp'],
            'mo_sales': ek['mo_sales'], 'mo_favs': ek['mo_favs'], 'mo_revs': ek['mo_revs'],
            'sdr': round(max(ek['mo_views'], sk['mo_views']) / ((sk['mo_comp'] if sk['mo_comp'] > 0 else ek['mo_comp']) + 1), 4),
            'cvr': sk['cvr'] if sk['cvr'] > 0 else ek['cvr'],
            'biz': ek['biz'], 'cpc': sk['cpc'],
            'position_vol': sk['position_vol'], 'top3_click': sk['top3_click'],
            'top3_cvr': sk['top3_cvr'], 'comp_count': sk['comp_count'],
            'rel_score': sk['rel_score'], 'rel_tier': sk['rel_tier'],
            'tier': sk['tier'], 'tag': sk['tag'],
            'decision': f'双源验证 | {sk["decision"]}',
            'source': '双源'
        }
    elif sk:
        d = sk
        d['decision'] = f'竞品独家 | {d["decision"]}'
        d['source'] = '西柚独有'
    else:
        d = ek
        d['tier'] = 'T5'
        d['tag'] = '待语义审查'
        d['decision'] = f'品类发现(待语义审查) | 流量={d["mo_views"]} | 待判定:污染隔离/待归类确认/已审定'
        d['source'] = 'erank独有'
        d['cpc'] = 0
        d['position_vol'] = 0
        d['top3_click'] = 0
        d['top3_cvr'] = 0
        d['comp_count'] = 0
        d['rel_score'] = 0
        d['rel_tier'] = '未验证'
    merged.append(d)

overlap = len([d for d in merged if d['source'] == '双源'])
s_only = len([d for d in merged if d['source'] == '西柚独有'])
e_only = len([d for d in merged if d['source'] == 'erank独有'])
tc2 = Counter(d['tier'] for d in merged)
print(f'\n融合: {len(merged)}词 (双源={overlap} 西柚独有={s_only} erank独有={e_only})')
print(f'融合T1-T5: {dict(tc2)}')

# ──── 4. 输出 ────
safe_name = SPU_NAME.replace("/", "_").replace("\\", "_").replace(":", "_")
OUT = os.path.join(BASE, '已处理关键词', f'{SPU_ID} {safe_name}.csv')

# 新版字段：基础 + 竞争情报 + 来源标记
H = ['关键词', '月浏览量', '竞争度', '供需比', '转化率', '商业价值(CPCxCVR)',
     'CPC建议竞价', '竞品数量', '排名滚动率', 'Top3点击垄断度',
     '相关性得分', '相关性档位',
     '词级', '词级标签', '决策分析', '数据源']

with open(OUT, 'w', encoding='utf-8-sig', newline='') as f:
    w = csv.writer(f, quoting=csv.QUOTE_ALL)
    w.writerow(H)
    for d in sorted(merged, key=lambda x: -x['mo_views']):
        w.writerow([
            d['kw'], d['mo_views'], d['mo_comp'], f'{d["sdr"]:.4f}',
            f'{d["cvr"] * 100:.2f}%' if isinstance(d['cvr'], float) and d['cvr'] < 1 else d['cvr'],
            f'{d["biz"]:.6f}',
            f'{d["cpc"]:.2f}' if d['cpc'] else '0.00',
            d['comp_count'],
            f'{d["position_vol"]:.2f}' if d['position_vol'] else '0.00',
            f'{d["top3_click"]:.4f}' if d['top3_click'] else '0.0000',
            d['rel_score'], d.get('rel_tier', ''),
            d['tier'], d['tag'], d['decision'], d['source']
        ])
print(f'CSV: {OUT} ({len(merged)}条)')

# ──── 4b. erank独有词独立待审查清单 ────
review_rows = [d for d in merged if d['source'] == 'erank独有']
if review_rows:
    REV = os.path.join(BASE, '已处理关键词', f'{SPU_ID} {safe_name}_待语义审查.csv')
    with open(REV, 'w', encoding='utf-8-sig', newline='') as rf:
        rw = csv.writer(rf, quoting=csv.QUOTE_ALL)
        rw.writerow(['关键词', '月浏览量', '竞争度', '供需比', '决策分析', '待判定'])
        for d in sorted(review_rows, key=lambda x: -x['mo_views']):
            rw.writerow([d['kw'], d['mo_views'], d['mo_comp'], f'{d["sdr"]:.4f}',
                         d['decision'], '污染隔离/待归类确认/已审定'])
    print(f'待审查清单: {REV} ({len(review_rows)}条)')

# ──── 5. SQLite ────
DB = os.path.join(BASE, 'keyword_database.db')
conn = sqlite3.connect(DB)
cur = conn.cursor()
dm = datetime.now().strftime("%Y-%m")
cur.execute("DELETE FROM keyword_tiers WHERE keyword_id IN (SELECT id FROM keywords WHERE spu_id=?)", (SPU_ID,))
cur.execute("DELETE FROM keywords WHERE spu_id=?", (SPU_ID,))
cur.execute("INSERT OR IGNORE INTO spu (spu_id,spu_name,category) VALUES (?,?,?)", (SPU_ID, SPU_NAME, CAT))
ok = 0
for d in merged:
    cvr_v = d['cvr'] if isinstance(d['cvr'], float) else 0
    try:
        cur.execute("INSERT INTO keywords (spu_id,keyword,competition,monthly_views,monthly_favs,monthly_sales,monthly_reviews,supply_demand_ratio,conversion_rate,purchase_intent,data_month) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    (SPU_ID, d['kw'], d['mo_comp'], d['mo_views'], d['mo_favs'], d['mo_sales'],
                     d['mo_revs'], d['sdr'], cvr_v, d['biz'], dm))
        cur.execute("INSERT INTO keyword_tiers (keyword_id,tier,tier_label,supply_demand_ratio,conversion_rate,purchase_intent) VALUES (?,?,?,?,?,?)",
                    (cur.lastrowid, d['tier'], d['tag'], d['sdr'], cvr_v, d['biz']))
        ok += 1
    except:
        pass
conn.commit()
cur.execute("SELECT COUNT(*) FROM keywords")
print(f'SQLite: {cur.fetchone()[0]} total, {ok} written')
conn.close()

# ──── 5b. 品类原始词池（永久保留，供同品类新SPU复用）────
try:
    from build_category_pool import upsert_spu
    n_pool = upsert_spu(SPU_ID, CAT, dm)
    print(f'品类词池: {n_pool} 词已吸入 (category={CAT})')
except Exception as e:
    print(f'品类词池更新失败(非致命): {e}')

# ──── 6. xlsx ────
try:
    import openpyxl

    xlsx_path = os.path.join(BASE, 'erank获取关键词.xlsx')
    wb = openpyxl.load_workbook(xlsx_path)
    for sn in ['关键词库', '词级词库']:
        ws = wb[sn]
        ir = ws.max_row + 1
        for i, d in enumerate(data_rows := merged):
            r = ir + i
            ws.cell(r, 1, SPU_ID)
            ws.cell(r, 2, SPU_NAME)
            ws.cell(r, 3, d['kw'])
            ws.cell(r, 4, d['mo_views'])
            ws.cell(r, 5, d['mo_comp'])
            ws.cell(r, 6, f'{d["sdr"]:.4f}')
            ws.cell(r, 7, f'{d["cvr"]*100:.2f}%' if isinstance(d['cvr'], float) and d['cvr']<1 else d['cvr'])
            ws.cell(r, 8, f'{d["biz"]:.6f}')
            ws.cell(r, 9, f'{d["cpc"]:.2f}' if d['cpc'] else '0.00')
            ws.cell(r, 9, d['comp_count'])
            ws.cell(r, 10, f'{d["position_vol"]:.2f}' if d['position_vol'] else '0.00')
            ws.cell(r, 11, f'{d["top3_click"]:.4f}' if d['top3_click'] else '0.0000')
            ws.cell(r, 12, d['rel_score'])
            ws.cell(r, 13, d.get('rel_tier', ''))
            ws.cell(r, 14, d['tier'])
            ws.cell(r, 15, d['tag'])
            ws.cell(r, 16, d['decision'])
            ws.cell(r, 17, d['source'])
            if sn == '词级词库':
                ws.cell(r, 18, d['tier'])
                ws.cell(r, 19, d['tag'])
    wb.save(xlsx_path)
    print(f'xlsx: {xlsx_path} updated')
except Exception as e:
    print(f'xlsx: skip ({e})')

# ──── 7. GitHub 同步 ────
try:
    sops_path = r"C:/Users/Administrator.DESKTOP-AHRMISP/Desktop/multi-agent-sop"
    if os.path.isdir(sops_path):
        import shutil
        shutil.copy2(DB, os.path.join(sops_path, 'keyword_database.db'))
        os.system(
            f'git -C "{sops_path}" add keyword_database.db && git -C "{sops_path}" commit -m "sync: {SPU_ID} {SPU_NAME}" && git -C "{sops_path}" push origin master')
        print('GitHub: pushed')
    else:
        print('GitHub: multi-agent-sop not found')
except Exception as e:
    print(f'GitHub: skip ({e})')

print(f'\n✅ {SPU_ID} {SPU_NAME} 完成')